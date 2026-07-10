"""Excel file reading functionality."""

import re
from pathlib import Path

import openpyxl

from models import EmployeeData, SubjectData, SubjectTitle


def _load_cell_value(file: Path, cell: str) -> str:
    """Load a single cell value from Excel file."""
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook.active

    if sheet is None:
        return ""

    cell_obj = sheet[cell]

    if cell_obj is None:
        return ""

    value = cell_obj.value

    return str(value) if value is not None else ""


def _parse_subject_abbreviation(text: str) -> str | None:
    """Extract abbreviation from subject title."""
    pattern = re.compile(r"\(([A-Za-z]{2,})\)$")
    match = pattern.search(text)

    if match:
        return match.group(1).lower()

    return None


def _parse_subject_title(text: str) -> str:
    """Extract title only from subject title, removing abbreviation."""
    pattern = re.compile(r"\s*\([A-Za-z]{2,}\)$")
    return pattern.sub("", text).replace("\n", " ").strip()


def load_employee_data(file: Path, cell: str) -> EmployeeData:
    """Load employee (candidate or assessor) data."""
    value = _load_cell_value(file, cell).lower()
    parts = value.split()

    if len(parts) < 2:
        return {
            "name": "",
            "license": "",
        }

    return {
        "name": " ".join(parts[:-1]).title(),
        "license": parts[-1].upper(),
    }


def load_subject_titles(file: Path, cell_range: str) -> list[SubjectTitle]:
    """Load subject titles and abbreviations."""
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook.active

    if sheet is None:
        return []

    subjects = []

    for row in sheet[cell_range]:
        cell_value = str(row[0].value or "")
        abbrev = _parse_subject_abbreviation(cell_value)
        title = _parse_subject_title(cell_value)

        if abbrev:
            subjects.append(SubjectTitle(abbreviation=abbrev, title=title))

    return subjects


def _load_numeric_values(file: Path, range_start: str, range_end: str) -> list[int]:
    """Load numeric values from a column range."""
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook.active
    column_range = f"{range_start}:{range_end}"

    if sheet is None:
        return []

    values = []

    for row in sheet[column_range]:
        cell_value = str(row[0].value or "")

        if cell_value.isdigit():
            values.append(int(cell_value))

    return values


def load_total_questions(file: Path, cell_range: str) -> list[int]:
    """Load total number of questions for each subject."""
    start, end = cell_range.split(":")
    return _load_numeric_values(file, start, end)


def load_percentages(file: Path, cell_range: str) -> list[int]:
    """Load percentage values for each subject."""
    start, end = cell_range.split(":")
    return _load_numeric_values(file, start, end)


def load_generated_numbers(file: Path, cell_range: str) -> list[list[int]]:
    """Load generated question numbers for each subject."""
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook.active
    pattern = re.compile(r"(\d{1,3},?\s*)+")

    if sheet is None:
        return []

    generated_numbers = []

    for row in sheet[cell_range]:
        cell_value = str(row[0].value or "")

        if pattern.search(cell_value):
            numbers = [int(n) for n in cell_value.split(", ")]
            generated_numbers.append(numbers)

    return generated_numbers


def load_all_subject_data(
    file: Path,
    subject_titles_cell_range: str,
    total_questions_cell_range: str,
    percentages_cell_range: str,
    generated_numbers_cell_range: str,
) -> list[SubjectData]:
    """Load complete data for each subject."""
    subject_titles = load_subject_titles(file, subject_titles_cell_range)
    total_questions = load_total_questions(file, total_questions_cell_range)
    percentages = load_percentages(file, percentages_cell_range)
    generated_numbers = load_generated_numbers(file, generated_numbers_cell_range)

    subjects: list[SubjectData] = []

    for i, subject_title in enumerate(subject_titles):
        subject: SubjectData = {
            "abbreviation": subject_title["abbreviation"],
            "title": subject_title["title"],
            "total_questions": total_questions[i],
            "percentage": percentages[i],
            "generated_numbers": generated_numbers[i],
        }
        subjects.append(subject)

    return subjects
